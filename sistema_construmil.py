# import tkinter as tk
# from tkinter import filedialog, messagebox, scrolledtext, ttk
# import pandas as pd
# import re
# import os
# import itertools
# import threading
# from pathlib import Path
# import warnings
# from openpyxl.styles import PatternFill, Font
#
# # Import pdfplumber (Exclusive for Amanco Engine)
# try:
#     import pdfplumber
# except ImportError:
#     pdfplumber = None
#
# warnings.filterwarnings('ignore')
#
#
# # ==============================================================================
# # BRAIN 1: DECA ENGINE (Logic provided by you - UNTOUCHED)
# # ==============================================================================
#
# class DecaEngine:
#     def normalizar_codigo_pdf(self, codigo: str) -> list:
#         if not codigo: return []
#         codigo = str(codigo).upper().strip()
#         variacoes = set([codigo, codigo.replace('.', ''), codigo.replace('.', '-'), codigo.replace('.', ' ')])
#
#         mapa_fracoes = {'012': '1/2', '034': '3/4', '001': '1', '014': '1/4', '112': '1 1/2', '114': '1 1/4'}
#         partes = codigo.split('.')
#
#         if len(partes) >= 2:
#             tokens = []
#             for p in partes:
#                 if p in mapa_fracoes:
#                     tokens.append([p, mapa_fracoes[p]])
#                 elif re.match(r'^[A-Z]\d{2,3}$', p):
#                     tokens.append([p, f"{p[0]}-{p[1:]}", f"{p[0]} {p[1:]}"])
#                 else:
#                     tokens.append([p])
#
#             for combo in itertools.product(*tokens):
#                 variacoes.update([" ".join(combo), "-".join(combo), "".join(combo), " ".join(reversed(combo))])
#                 if len(combo) == 3:
#                     variacoes.update([" ".join(p) for p in itertools.permutations(combo)])
#                     variacoes.add(f"{combo[0]}{combo[1]} {combo[2]}")
#                     variacoes.add(f"{combo[0]}{combo[1]} {combo[2].replace('/', '')}")
#
#         return [v.strip() for v in variacoes if len(v.strip()) > 2]
#
#     def buscar_codigo_na_descricao(self, descricao: str, variacoes_codigo: list) -> bool:
#         if not descricao or not variacoes_codigo: return False
#         desc_upper = str(descricao).upper()
#         desc_super_limpa = re.sub(r'[^A-Z0-9]', '', desc_upper)
#
#         for v in variacoes_codigo:
#             v_upper = v.upper()
#             if v_upper in desc_upper: return True
#             tokens = v_upper.split()
#             if len(tokens) > 1:
#                 desc_limpa_tokens = re.sub(r'[^\w\s/]', ' ', desc_upper)
#                 if all((t in desc_limpa_tokens or t in desc_upper) for t in tokens if len(t) > 1):
#                     return True
#             v_super_limpa = re.sub(r'[^A-Z0-9]', '', v_upper)
#             if len(v_super_limpa) > 4 and v_super_limpa in desc_super_limpa:
#                 return True
#         return False
#
#     def limpar_valor(self, valor) -> float:
#         if pd.isna(valor) or valor == '': return 0.0
#         v = str(valor).replace('R$', '').replace('BRL', '').strip()
#         if re.search(r'\d{1,3}(\.\d{3})*,\d{2}', v):
#             v = v.replace('.', '').replace(',', '.')
#         elif ',' in v and v.count(',') == 1 and len(v.split(',')[1]) == 2:
#             v = v.replace(',', '.')
#         try:
#             return float(re.sub(r'[^\d.]', '', v))
#         except:
#             return 0.0
#
#     def ler_pdf_extraido(self, caminho_txt: str) -> pd.DataFrame:
#         with open(caminho_txt, 'r', encoding='utf-8') as f:
#             conteudo = f.read()
#         produtos = []
#         padroes = [
#             r'(?m)^([A-Z0-9][A-Z0-9.\-]{2,}(?:\.[A-Z0-9]+)*)\s+(?:.*?)\s+(\d{2}/\d{2}/\d{4})\s+(?:BRL\s*)?([\d.,]+)\s+([\d.,]+)',
#             r'(?m)^([A-Z0-9][A-Z0-9.\-]{2,}(?:\.[A-Z0-9]+)*)\t.*?\t(\d{2}/\d{2}/\d{4})\t(?:BRL\s*)?([\d.,]+)\t([\d.,]+)'
#         ]
#         for padrao in padroes:
#             for match in re.findall(padrao, conteudo):
#                 if len(match) < 4: continue
#                 cod, data, val_s, qtd_s = match[:4]
#                 cod = cod.strip()
#                 if len(cod) < 5 or cod == '2025' or 'BRL' in cod or '/' in cod: continue
#                 qtd = int(self.limpar_valor(qtd_s))
#                 if qtd > 0: produtos.append({'Código PDF': cod, 'Qtd PDF': qtd, 'Valor PDF': self.limpar_valor(val_s)})
#         df = pd.DataFrame(produtos)
#         return df.drop_duplicates(subset=['Código PDF']) if not df.empty else df
#
#     def ler_excel(self, caminho_excel: str) -> pd.DataFrame:
#         df = None
#         if not caminho_excel.lower().endswith('.csv'):
#             try:
#                 df = pd.read_excel(caminho_excel, sheet_name="Produtos")
#             except:
#                 try:
#                     df = pd.read_excel(caminho_excel)
#                 except:
#                     pass
#         if df is None:
#             for enc in ['utf-8', 'latin1']:
#                 for sep in [';', ',']:
#                     try:
#                         temp = pd.read_csv(caminho_excel, encoding=enc, sep=sep)
#                         if temp.shape[1] > 1: df = temp; break
#                     except:
#                         continue
#                 if df is not None: break
#         if df is None: raise Exception("Falha ao ler Excel/CSV.")
#         df.columns = [str(c).strip() for c in df.columns]
#
#         col_prod = next((c for c in df.columns if 'descri' in c.lower()), None)
#         if not col_prod:
#             for c in df.columns:
#                 cl = c.lower()
#                 if ('produto' in cl or 'nome' in cl) and not ('código' in cl or 'codigo' in cl or 'cod' in cl):
#                     col_prod = c;
#                     break
#         if not col_prod:
#             col_prod = next((c for c in df.columns if 'produto' in c.lower()), None)
#         col_qtd = next((c for c in df.columns if any(x in c.lower() for x in ['compra', 'qtd', 'quantidade'])), None)
#         col_val = next((c for c in df.columns if any(x in c.lower() for x in ['valor', 'preço', 'vlr', 'preco'])), None)
#
#         if not col_prod: raise Exception(f"Coluna de DESCRIÇÃO não encontrada.")
#         df_novo = pd.DataFrame()
#         df_novo['Produto Excel'] = df[col_prod]
#         df_novo['Qtd Excel'] = df[col_qtd] if col_qtd else 0
#         df_novo['Valor Excel'] = df[col_val] if col_val else 0.0
#         return df_novo
#
#     def processar(self, txt_path, excel_path, log_func):
#         log_func("DECA: Lendo PDF (TXT)...")
#         df_pdf = self.ler_pdf_extraido(txt_path)
#         log_func(f"DECA: {len(df_pdf)} códigos no PDF.")
#
#         log_func("DECA: Lendo Excel...")
#         df_excel = self.ler_excel(excel_path)
#         log_func(f"DECA: {len(df_excel)} linhas no Excel.")
#
#         log_func("DECA: Cruzando dados...")
#         resultados = []
#         matches = 0
#
#         for i, row in df_pdf.iterrows():
#             cod, qtd, val = str(row['Código PDF']), row['Qtd PDF'], row['Valor PDF']
#             variacoes = self.normalizar_codigo_pdf(cod)
#             match_data = None
#             melhor_diff = float('inf')
#
#             for _, row_ex in df_excel.iterrows():
#                 desc = str(row_ex['Produto Excel'])
#                 if self.buscar_codigo_na_descricao(desc, variacoes):
#                     diff = abs(val - self.limpar_valor(row_ex['Valor Excel']))
#                     if diff < melhor_diff:
#                         melhor_diff = diff
#                         match_data = row_ex
#
#             diff_qtd_final = qtd
#             diff_val_final = val
#             status = "❌ NÃO ENCONTRADO"
#             prod_excel = "---"
#             qtd_excel = 0
#             val_excel = 0.0
#
#             if match_data is not None:
#                 matches += 1
#                 qtd_excel = int(self.limpar_valor(match_data['Qtd Excel']))
#                 val_excel = self.limpar_valor(match_data['Valor Excel'])
#                 diff_qtd_final = qtd - qtd_excel
#                 diff_val_final = val - val_excel
#                 prod_excel = match_data['Produto Excel']
#                 status = "✓ OK" if (diff_qtd_final == 0) and (abs(diff_val_final) < 0.2) else "⚠ DIVERGENTE"
#
#             resultados.append({
#                 'Código PDF': cod, 'Produto Excel': prod_excel,
#                 'Qtd PDF': qtd, 'Qtd Excel': qtd_excel, 'Diferença Qtd': diff_qtd_final,
#                 'Valor PDF': val, 'Valor Excel': val_excel, 'Diferença Valor': diff_val_final,
#                 'Status': status
#             })
#
#         log_func(f"DECA: Matches: {matches}/{len(df_pdf)}")
#
#         # Save Deca
#         out_path = Path(txt_path).parent / (Path(txt_path).stem + "_DECA_RELATORIO.xlsx")
#         df_res = pd.DataFrame(resultados)
#         with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
#             df_res.to_excel(writer, index=False, sheet_name='Conferência')
#             ws = writer.sheets['Conferência']
#             red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
#             green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
#             for row in range(2, ws.max_row + 1):
#                 c = ws.cell(row=row, column=5)  # Diff Qtd
#                 if c.value and c.value != 0: c.fill = red if c.value < 0 else green
#                 c = ws.cell(row=row, column=8)  # Diff Val
#                 if c.value and c.value != 0: c.fill = red if c.value > 0 else green
#             for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 15
#
#         return out_path
#
#
# # ==============================================================================
# # BRAIN 2: AMANCO ENGINE (Logic provided by you - UNTOUCHED)
# # ==============================================================================
#
# class AmancoEngine:
#     class PriceComparator:
#         def __init__(self, excel_path, pdf_path, logger=None):
#             self.excel_path = excel_path
#             self.pdf_path = pdf_path
#             self.df_excel = None
#             self.df_pdf = None
#             self.logger = logger  # Function to redirect print to GUI
#
#         def log(self, msg):
#             if self.logger:
#                 self.logger(msg)
#             else:
#                 print(msg)
#
#         def extract_product_code(self, product_text):
#             if pd.isna(product_text): return None
#             matches = re.findall(r'\b(\d{4,6})\b', str(product_text))
#             return matches[-1] if matches else None
#
#         def clean_price(self, price_text):
#             if pd.isna(price_text): return None
#             price_str = str(price_text).replace('R$', '').replace('.', '').replace(',', '.').strip()
#             try:
#                 return float(price_str)
#             except ValueError:
#                 return None
#
#         def read_excel(self):
#             self.log(f"Lendo arquivo Excel: {Path(self.excel_path).name}")
#             try:
#                 self.df_excel = pd.read_excel(self.excel_path)
#                 self.df_excel['Código Extraído'] = self.df_excel['Produto'].apply(self.extract_product_code)
#                 if 'Compra' in self.df_excel.columns:
#                     self.df_excel['Compra'] = pd.to_numeric(self.df_excel['Compra'], errors='coerce')
#                 if 'Valor de compra' in self.df_excel.columns:
#                     self.df_excel['Valor de compra'] = pd.to_numeric(self.df_excel['Valor de compra'], errors='coerce')
#                 self.log(f"Excel lido com sucesso: {len(self.df_excel)} linhas")
#             except Exception as e:
#                 self.log(f"Erro ao ler Excel: {e}")
#                 raise
#
#         def read_pdf(self):
#             self.log(f"Lendo arquivo PDF: {Path(self.pdf_path).name}")
#             if pdfplumber is None: raise Exception("pdfplumber não instalado.")
#             try:
#                 pdf_data = []
#                 with pdfplumber.open(self.pdf_path) as pdf:
#                     self.log(f"Total de páginas: {len(pdf.pages)}")
#                     for page_num, page in enumerate(pdf.pages, 1):
#                         tables = page.extract_tables()
#                         if tables:
#                             for table_idx, table in enumerate(tables):
#                                 if not table or len(table) < 2: continue
#                                 header = [str(h).lower() if h else '' for h in table[0]]
#                                 codigo_idx = next((i for i, h in enumerate(header) if 'código' in h or 'codigo' in h),
#                                                   None)
#                                 qtde_idx = next((i for i, h in enumerate(header) if 'qtde' in h or 'quantidade' in h),
#                                                 None)
#                                 preco_idx = next((i for i, h in enumerate(header) if 'preço' in h or 'preco' in h),
#                                                  None)
#                                 if None in (codigo_idx, qtde_idx, preco_idx): continue
#                                 for row in table[1:]:
#                                     try:
#                                         if len(row) > max(codigo_idx, qtde_idx, preco_idx):
#                                             c = str(row[codigo_idx]).strip() if row[codigo_idx] else None
#                                             if c and re.match(r'^\d{4,6}$', c):
#                                                 pdf_data.append({'Código_PDF': c, 'Qtde_PDF': row[qtde_idx],
#                                                                  'Preço_Líq_PDF': row[preco_idx]})
#                                     except:
#                                         continue
#                         if not tables or not pdf_data:
#                             text = page.extract_text()
#                             if text:
#                                 for line in text.split('\n'):
#                                     match = re.match(
#                                         r'^(\d{4,6})\s+.*?(?:BR\d+|0\d+)\s*-\s*.*?\s+(\d+)\s+R\$\s*([\d.,]+)', line)
#                                     if match:
#                                         pdf_data.append({'Código_PDF': match.group(1), 'Qtde_PDF': match.group(2),
#                                                          'Preço_Líq_PDF': match.group(3)})
#                 self.df_pdf = pd.DataFrame(pdf_data)
#                 if not self.df_pdf.empty:
#                     self.df_pdf['Qtde_PDF'] = pd.to_numeric(self.df_pdf['Qtde_PDF'], errors='coerce')
#                     self.df_pdf['Preço_Líq_PDF'] = self.df_pdf['Preço_Líq_PDF'].apply(self.clean_price)
#                     self.df_pdf = self.df_pdf.drop_duplicates(subset=['Código_PDF'], keep='first')
#                 self.log(f"PDF lido com sucesso: {len(self.df_pdf)} itens")
#             except Exception as e:
#                 self.log(f"Erro ao ler PDF: {e}")
#                 raise
#
#         def merge_data(self):
#             self.log("Combinando dados...")
#             if self.df_pdf.empty:
#                 df_result = self.df_excel.copy()
#                 df_result['Qtde_PDF'] = None;
#                 df_result['Preço_Líq_PDF'] = None
#             else:
#                 df_result = self.df_excel.merge(self.df_pdf, left_on='Código Extraído', right_on='Código_PDF',
#                                                 how='left')
#                 if 'Código_PDF' in df_result.columns: df_result = df_result.drop('Código_PDF', axis=1)
#             matches = df_result['Qtde_PDF'].notna().sum()
#             self.log(f"Combinação concluída: {matches} correspondências")
#             return df_result
#
#         def calculate_differences(self, df_result):
#             self.log("Calculando diferenças...")
#             df_result['Diferença de Qtde'] = df_result['Qtde_PDF'] - df_result['Compra']
#             df_result['Diferença de Preço'] = df_result['Preço_Líq_PDF'] - df_result['Valor de compra']
#             return df_result
#
#         def save_result(self, df_result, output_path):
#             self.log(f"Salvando em: {Path(output_path).name}")
#             with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#                 df_result.to_excel(writer, index=False, sheet_name='Comparação')
#                 ws = writer.sheets['Comparação']
#                 red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
#                 green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
#                 bold = Font(bold=True)
#                 for cell in ws[1]: cell.font = bold
#                 headers = [c.value for c in ws[1]]
#                 try:
#                     qc = headers.index('Diferença de Qtde') + 1
#                     pc = headers.index('Diferença de Preço') + 1
#                     for row in range(2, len(df_result) + 2):
#                         c = ws.cell(row=row, column=qc)
#                         if c.value:
#                             if float(c.value) > 0:
#                                 c.fill = green
#                             elif float(c.value) < 0:
#                                 c.fill = red
#                         c = ws.cell(row=row, column=pc)
#                         if c.value:
#                             if float(c.value) > 0:
#                                 c.fill = red
#                             elif float(c.value) < 0:
#                                 c.fill = green
#                 except:
#                     pass
#                 for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
#
#     def processar(self, pdf_path, excel_path, log_func):
#         out_path = Path(pdf_path).parent / (Path(pdf_path).stem + "_AMANCO_RELATORIO.xlsx")
#         # Wrapper to instantiate and run the PriceComparator logic provided
#         comparator = self.PriceComparator(excel_path, pdf_path, log_func)
#         comparator.read_excel()
#         comparator.read_pdf()
#         df = comparator.merge_data()
#         df = comparator.calculate_differences(df)
#         comparator.save_result(df, str(out_path))
#         return out_path
#
#
# # ==============================================================================
# # UNIFIED INTERFACE (BLUE/RED THEME)
# # ==============================================================================
#
# class ConstrumilApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("SISTEMA CONSTRUMIL - Conferência v5.0")
#         self.root.geometry("900x700")
#         self.bg, self.acc, self.pnl = "#0A1931", "#D32F2F", "#182B45"
#         self.root.configure(bg=self.bg)
#
#         # Header
#         hdr = tk.Frame(root, bg=self.bg)
#         hdr.pack(fill="x", pady=20)
#         tk.Label(hdr, text="CONSTRUMIL", font=("Arial Black", 28), bg=self.bg, fg="white").pack()
#         tk.Label(hdr, text="Sistema de Conferência Inteligente", font=("Arial", 10), bg=self.bg, fg="#AAA").pack()
#         tk.Frame(root, bg=self.acc, height=5).pack(fill="x")
#
#         # Tabs
#         style = ttk.Style()
#         style.theme_use('clam')
#         style.configure("TNotebook", background=self.bg, borderwidth=0)
#         style.configure("TNotebook.Tab", background="#333", foreground="white", padding=[20, 10],
#                         font=('Arial', 10, 'bold'))
#         style.map("TNotebook.Tab", background=[("selected", self.acc)], foreground=[("selected", "white")])
#
#         self.nb = ttk.Notebook(root)
#         self.nb.pack(fill="both", expand=True, padx=20, pady=20)
#
#         self.tab_deca = tk.Frame(self.nb, bg=self.bg)
#         self.tab_amanco = tk.Frame(self.nb, bg=self.bg)
#
#         self.nb.add(self.tab_deca, text="  DECA  ")
#         self.nb.add(self.tab_amanco, text=" AMANCO ")
#
#         self.setup_ui(self.tab_deca, "DECA", "txt")
#         self.setup_ui(self.tab_amanco, "AMANCO", "pdf")
#
#         tk.Label(root, text="Construmil © 2025", bg=self.bg, fg="#555").pack(pady=5)
#
#     def setup_ui(self, parent, marca, ext):
#         card = tk.Frame(parent, bg=self.pnl, padx=20, pady=20)
#         card.pack(fill="both", expand=True)
#
#         setattr(self, f"in_{marca}", tk.StringVar())
#         setattr(self, f"xls_{marca}", tk.StringVar())
#
#         lbl = "Arquivo TXT (Extraído):" if ext == "txt" else "Arquivo PDF (Original):"
#         self.mk_input(card, lbl, getattr(self, f"in_{marca}"), ext)
#         self.mk_input(card, "Planilha Excel:", getattr(self, f"xls_{marca}"), "xls")
#
#         btn = tk.Button(card, text=f"PROCESSAR {marca}", bg=self.acc, fg="white",
#                         font=("Arial", 12, "bold"), relief="flat", cursor="hand2",
#                         command=lambda: self.run_process(marca))
#         btn.pack(fill="x", pady=20, ipady=5)
#
#         tk.Label(card, text="Log:", bg=self.pnl, fg="white").pack(anchor="w")
#         log = scrolledtext.ScrolledText(card, height=12, bg="#0F1F30", fg="#0F0", font=("Consolas", 9))
#         log.pack(fill="both", expand=True)
#         setattr(self, f"log_{marca}", log)
#
#     def mk_input(self, parent, text, var, ext):
#         f = tk.Frame(parent, bg=self.pnl)
#         f.pack(fill="x", pady=5)
#         tk.Label(f, text=text, bg=self.pnl, fg="#CCC", width=25, anchor="w").pack(side="left")
#         tk.Entry(f, textvariable=var).pack(side="left", fill="x", expand=True, padx=5)
#         tk.Button(f, text="📂", command=lambda: self.browse(var, ext)).pack(side="left")
#
#     def browse(self, var, ext):
#         ft = [("PDF", "*.pdf")] if ext == "pdf" else [("Text", "*.txt")] if ext == "txt" else [
#             ("Excel", "*.xlsx *.xls")]
#         f = filedialog.askopenfilename(filetypes=ft)
#         if f: var.set(f)
#
#     def log(self, marca, msg):
#         l = getattr(self, f"log_{marca}")
#         l.insert(tk.END, f"> {msg}\n")
#         l.see(tk.END)
#
#     def run_process(self, marca):
#         inp = getattr(self, f"in_{marca}").get()
#         xls = getattr(self, f"xls_{marca}").get()
#         if not inp or not xls:
#             messagebox.showwarning("Erro", "Selecione os arquivos!")
#             return
#         threading.Thread(target=lambda: self.worker(marca, inp, xls), daemon=True).start()
#
#     def worker(self, marca, inp, xls):
#         try:
#             self.log(marca, f"Iniciando Engine {marca}...")
#             out_path = ""
#
#             if marca == "DECA":
#                 engine = DecaEngine()
#                 out_path = engine.processar(inp, xls, lambda m: self.log(marca, m))
#             else:
#                 engine = AmancoEngine()
#                 out_path = engine.processar(inp, xls, lambda m: self.log(marca, m))
#
#             self.log(marca, f"SUCESSO! Salvo.")
#             messagebox.showinfo("Sucesso", f"Processamento concluído!\n\nSalvo em:\n{out_path}")
#
#         except Exception as e:
#             self.log(marca, f"ERRO FATAL: {e}")
#             messagebox.showerror("Erro", str(e))
#
#
# if __name__ == "__main__":
#     root = tk.Tk()
#     ConstrumilApp(root)
#     root.mainloop()


import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import pandas as pd
import re
import os
import itertools
import threading
from pathlib import Path
import warnings
from openpyxl.styles import PatternFill, Font
# Importação necessária para imagens PNG
from PIL import Image, ImageTk

# Tenta importar pdfplumber (Essencial para Amanco)
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

warnings.filterwarnings('ignore')


# ==============================================================================
# MOTOR 1: DECA ENGINE (Lógica Intocada)
# ==============================================================================

class DecaEngine:
    def normalizar_codigo_pdf(self, codigo: str) -> list:
        if not codigo: return []
        codigo = str(codigo).upper().strip()
        variacoes = set([codigo, codigo.replace('.', ''), codigo.replace('.', '-'), codigo.replace('.', ' ')])

        mapa_fracoes = {'012': '1/2', '034': '3/4', '001': '1', '014': '1/4', '112': '1 1/2', '114': '1 1/4'}
        partes = codigo.split('.')

        if len(partes) >= 2:
            tokens = []
            for p in partes:
                if p in mapa_fracoes:
                    tokens.append([p, mapa_fracoes[p]])
                elif re.match(r'^[A-Z]\d{2,3}$', p):
                    tokens.append([p, f"{p[0]}-{p[1:]}", f"{p[0]} {p[1:]}"])
                else:
                    tokens.append([p])

            for combo in itertools.product(*tokens):
                variacoes.update([" ".join(combo), "-".join(combo), "".join(combo), " ".join(reversed(combo))])
                if len(combo) == 3:
                    variacoes.update([" ".join(p) for p in itertools.permutations(combo)])
                    variacoes.add(f"{combo[0]}{combo[1]} {combo[2]}")
                    variacoes.add(f"{combo[0]}{combo[1]} {combo[2].replace('/', '')}")

        return [v.strip() for v in variacoes if len(v.strip()) > 2]

    def buscar_codigo_na_descricao(self, descricao: str, variacoes_codigo: list) -> bool:
        if not descricao or not variacoes_codigo: return False
        desc_upper = str(descricao).upper()
        desc_super_limpa = re.sub(r'[^A-Z0-9]', '', desc_upper)

        for v in variacoes_codigo:
            v_upper = v.upper()
            if v_upper in desc_upper: return True
            tokens = v_upper.split()
            if len(tokens) > 1:
                desc_limpa_tokens = re.sub(r'[^\w\s/]', ' ', desc_upper)
                if all((t in desc_limpa_tokens or t in desc_upper) for t in tokens if len(t) > 1):
                    return True
            v_super_limpa = re.sub(r'[^A-Z0-9]', '', v_upper)
            if len(v_super_limpa) > 4 and v_super_limpa in desc_super_limpa:
                return True
        return False

    def limpar_valor(self, valor) -> float:
        if pd.isna(valor) or valor == '': return 0.0
        v = str(valor).replace('R$', '').replace('BRL', '').strip()
        if re.search(r'\d{1,3}(\.\d{3})*,\d{2}', v):
            v = v.replace('.', '').replace(',', '.')
        elif ',' in v and v.count(',') == 1 and len(v.split(',')[1]) == 2:
            v = v.replace(',', '.')
        try:
            return float(re.sub(r'[^\d.]', '', v))
        except:
            return 0.0

    def ler_pdf_extraido(self, caminho_txt: str) -> pd.DataFrame:
        with open(caminho_txt, 'r', encoding='utf-8') as f:
            conteudo = f.read()
        produtos = []
        padroes = [
            r'(?m)^([A-Z0-9][A-Z0-9.\-]{2,}(?:\.[A-Z0-9]+)*)\s+(?:.*?)\s+(\d{2}/\d{2}/\d{4})\s+(?:BRL\s*)?([\d.,]+)\s+([\d.,]+)',
            r'(?m)^([A-Z0-9][A-Z0-9.\-]{2,}(?:\.[A-Z0-9]+)*)\t.*?\t(\d{2}/\d{2}/\d{4})\t(?:BRL\s*)?([\d.,]+)\t([\d.,]+)'
        ]
        for padrao in padroes:
            for match in re.findall(padrao, conteudo):
                if len(match) < 4: continue
                cod, data, val_s, qtd_s = match[:4]
                cod = cod.strip()
                if len(cod) < 5 or cod == '2025' or 'BRL' in cod or '/' in cod: continue
                qtd = int(self.limpar_valor(qtd_s))
                if qtd > 0: produtos.append({'Código PDF': cod, 'Qtd PDF': qtd, 'Valor PDF': self.limpar_valor(val_s)})
        df = pd.DataFrame(produtos)
        return df.drop_duplicates(subset=['Código PDF']) if not df.empty else df

    def ler_excel(self, caminho_excel: str) -> pd.DataFrame:
        df = None
        if not caminho_excel.lower().endswith('.csv'):
            try:
                df = pd.read_excel(caminho_excel, sheet_name="Produtos")
            except:
                try:
                    df = pd.read_excel(caminho_excel)
                except:
                    pass
        if df is None:
            for enc in ['utf-8', 'latin1']:
                for sep in [';', ',']:
                    try:
                        temp = pd.read_csv(caminho_excel, encoding=enc, sep=sep)
                        if temp.shape[1] > 1: df = temp; break
                    except:
                        continue
                if df is not None: break
        if df is None: raise Exception("Falha ao ler Excel/CSV.")
        df.columns = [str(c).strip() for c in df.columns]

        col_prod = next((c for c in df.columns if 'descri' in c.lower()), None)
        if not col_prod:
            for c in df.columns:
                cl = c.lower()
                if ('produto' in cl or 'nome' in cl) and not ('código' in cl or 'codigo' in cl or 'cod' in cl):
                    col_prod = c;
                    break
        if not col_prod:
            col_prod = next((c for c in df.columns if 'produto' in c.lower()), None)
        col_qtd = next((c for c in df.columns if any(x in c.lower() for x in ['compra', 'qtd', 'quantidade'])), None)
        col_val = next((c for c in df.columns if any(x in c.lower() for x in ['valor', 'preço', 'vlr', 'preco'])), None)

        if not col_prod: raise Exception(f"Coluna de DESCRIÇÃO não encontrada.")
        df_novo = pd.DataFrame()
        df_novo['Produto Excel'] = df[col_prod]
        df_novo['Qtd Excel'] = df[col_qtd] if col_qtd else 0
        df_novo['Valor Excel'] = df[col_val] if col_val else 0.0
        return df_novo

    def processar(self, txt_path, excel_path, log_func):
        log_func("DECA: Lendo PDF (TXT)...")
        df_pdf = self.ler_pdf_extraido(txt_path)
        log_func(f"DECA: {len(df_pdf)} códigos no PDF.")

        log_func("DECA: Lendo Excel...")
        df_excel = self.ler_excel(excel_path)
        log_func(f"DECA: {len(df_excel)} linhas no Excel.")

        log_func("DECA: Cruzando dados...")
        resultados = []
        matches = 0

        for i, row in df_pdf.iterrows():
            cod, qtd, val = str(row['Código PDF']), row['Qtd PDF'], row['Valor PDF']
            variacoes = self.normalizar_codigo_pdf(cod)
            match_data = None
            melhor_diff = float('inf')

            for _, row_ex in df_excel.iterrows():
                desc = str(row_ex['Produto Excel'])
                if self.buscar_codigo_na_descricao(desc, variacoes):
                    diff = abs(val - self.limpar_valor(row_ex['Valor Excel']))
                    if diff < melhor_diff:
                        melhor_diff = diff
                        match_data = row_ex

            diff_qtd_final = qtd
            diff_val_final = val
            status = "❌ NÃO ENCONTRADO"
            prod_excel = "---"
            qtd_excel = 0
            val_excel = 0.0

            if match_data is not None:
                matches += 1
                qtd_excel = int(self.limpar_valor(match_data['Qtd Excel']))
                val_excel = self.limpar_valor(match_data['Valor Excel'])
                diff_qtd_final = qtd - qtd_excel
                diff_val_final = val - val_excel
                prod_excel = match_data['Produto Excel']
                status = "✓ OK" if (diff_qtd_final == 0) and (abs(diff_val_final) < 0.2) else "⚠ DIVERGENTE"

            resultados.append({
                'Código PDF': cod, 'Produto Excel': prod_excel,
                'Qtd PDF': qtd, 'Qtd Excel': qtd_excel, 'Diferença Qtd': diff_qtd_final,
                'Valor PDF': val, 'Valor Excel': val_excel, 'Diferença Valor': diff_val_final,
                'Status': status
            })

        log_func(f"DECA: Matches: {matches}/{len(df_pdf)}")

        out_path = Path(txt_path).parent / (Path(txt_path).stem + "_DECA_RELATORIO.xlsx")
        df_res = pd.DataFrame(resultados)
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df_res.to_excel(writer, index=False, sheet_name='Conferência')
            ws = writer.sheets['Conferência']
            red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            bold = Font(bold=True)
            for cell in ws[1]: cell.font = bold

            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=5)  # Diff Qtd
                if c.value and c.value != 0: c.fill = red if c.value < 0 else green
                c = ws.cell(row=row, column=8)  # Diff Val
                if c.value and c.value != 0: c.fill = red if c.value > 0 else green
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 15

        return out_path


# ==============================================================================
# MOTOR 2: AMANCO ENGINE (Lógica Intocada)
# ==============================================================================

class AmancoEngine:
    class PriceComparator:
        def __init__(self, excel_path, pdf_path, logger=None):
            self.excel_path = excel_path
            self.pdf_path = pdf_path
            self.df_excel = None
            self.df_pdf = None
            self.logger = logger

        def log(self, msg):
            if self.logger:
                self.logger(msg)
            else:
                print(msg)

        def extract_product_code(self, product_text):
            if pd.isna(product_text): return None
            matches = re.findall(r'\b(\d{4,6})\b', str(product_text))
            return matches[-1] if matches else None

        def clean_price(self, price_text):
            if pd.isna(price_text): return None
            price_str = str(price_text).replace('R$', '').replace('.', '').replace(',', '.').strip()
            try:
                return float(price_str)
            except ValueError:
                return None

        def read_excel(self):
            self.log(f"Lendo arquivo Excel: {Path(self.excel_path).name}")
            try:
                self.df_excel = pd.read_excel(self.excel_path)
                col_prod = next((c for c in self.df_excel.columns if 'produto' in c.lower()), 'Produto')
                self.df_excel['Código Extraído'] = self.df_excel[col_prod].apply(self.extract_product_code)
                if 'Compra' in self.df_excel.columns:
                    self.df_excel['Compra'] = pd.to_numeric(self.df_excel[c], errors='coerce')
                if 'Valor de compra' in self.df_excel.columns:
                    self.df_excel['Valor de compra'] = pd.to_numeric(self.df_excel[c], errors='coerce')
                self.log(f"Excel lido com sucesso: {len(self.df_excel)} linhas")
            except Exception as e:
                self.log(f"Erro Excel: {e}")
                raise

        def read_pdf(self):
            self.log(f"Lendo arquivo PDF: {Path(self.pdf_path).name}")
            if pdfplumber is None: raise Exception("pdfplumber não instalado.")
            try:
                pdf_data = []
                with pdfplumber.open(self.pdf_path) as pdf:
                    self.log(f"Total de páginas: {len(pdf.pages)}")
                    for page_num, page in enumerate(pdf.pages, 1):
                        tables = page.extract_tables()
                        if tables:
                            for table_idx, table in enumerate(tables):
                                if not table or len(table) < 2: continue
                                header = [str(h).lower() if h else '' for h in table[0]]
                                codigo_idx = next((i for i, h in enumerate(header) if 'código' in h or 'codigo' in h),
                                                  None)
                                qtde_idx = next((i for i, h in enumerate(header) if 'qtde' in h or 'quantidade' in h),
                                                None)
                                preco_idx = next((i for i, h in enumerate(header) if 'preço' in h or 'preco' in h),
                                                 None)
                                if None in (codigo_idx, qtde_idx, preco_idx): continue
                                for row in table[1:]:
                                    try:
                                        if len(row) > max(codigo_idx, qtde_idx, preco_idx):
                                            c = str(row[codigo_idx]).strip() if row[codigo_idx] else None
                                            if c and re.match(r'^\d{4,6}$', c):
                                                pdf_data.append({'Código_PDF': c, 'Qtde_PDF': row[qtde_idx],
                                                                 'Preço_Líq_PDF': row[preco_idx]})
                                    except:
                                        continue
                        if not tables or not pdf_data:
                            text = page.extract_text()
                            if text:
                                for line in text.split('\n'):
                                    match = re.match(
                                        r'^(\d{4,6})\s+.*?(?:BR\d+|0\d+)\s*-\s*.*?\s+(\d+)\s+R\$\s*([\d.,]+)', line)
                                    if match:
                                        pdf_data.append({'Código_PDF': match.group(1), 'Qtde_PDF': match.group(2),
                                                         'Preço_Líq_PDF': match.group(3)})
                self.df_pdf = pd.DataFrame(pdf_data)
                if not self.df_pdf.empty:
                    self.df_pdf['Qtde_PDF'] = pd.to_numeric(self.df_pdf['Qtde_PDF'], errors='coerce')
                    self.df_pdf['Preço_Líq_PDF'] = self.df_pdf['Preço_Líq_PDF'].apply(self.clean_price)
                    self.df_pdf = self.df_pdf.drop_duplicates(subset=['Código_PDF'], keep='first')
                self.log(f"PDF lido com sucesso: {len(self.df_pdf)} itens")
            except Exception as e:
                self.log(f"Erro PDF: {e}")
                raise

        def merge_data(self):
            self.log("Combinando dados...")
            if self.df_pdf.empty:
                df_result = self.df_excel.copy()
                df_result['Qtde_PDF'] = None;
                df_result['Preço_Líq_PDF'] = None
            else:
                df_result = self.df_excel.merge(self.df_pdf, left_on='Código Extraído', right_on='Código_PDF',
                                                how='left')
                if 'Código_PDF' in df_result.columns: df_result = df_result.drop('Código_PDF', axis=1)
            matches = df_result['Qtde_PDF'].notna().sum()
            self.log(f"Combinação concluída: {matches} correspondências")
            return df_result

        def calculate_differences(self, df_result):
            self.log("Calculando diferenças...")
            df_result['Diferença de Qtde'] = df_result['Qtde_PDF'] - df_result.get('Compra', 0)
            df_result['Diferença de Preço'] = df_result['Preço_Líq_PDF'] - df_result.get('Valor de compra', 0)
            return df_result

        def save_result(self, df_result, output_path):
            self.log(f"Salvando em: {Path(output_path).name}")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name='Comparação')
                ws = writer.sheets['Comparação']
                red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                bold = Font(bold=True)
                for cell in ws[1]: cell.font = bold
                headers = [c.value for c in ws[1]]
                try:
                    qc = headers.index('Diferença de Qtde') + 1
                    pc = headers.index('Diferença de Preço') + 1
                    for row in range(2, len(df_result) + 2):
                        c = ws.cell(row=row, column=qc)
                        if c.value:
                            if float(c.value) > 0:
                                c.fill = green
                            elif float(c.value) < 0:
                                c.fill = red
                        c = ws.cell(row=row, column=pc)
                        if c.value:
                            if float(c.value) > 0:
                                c.fill = red
                            elif float(c.value) < 0:
                                c.fill = green
                except:
                    pass
                for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20

    def processar(self, pdf_path, excel_path, log_func):
        out_path = Path(pdf_path).parent / (Path(pdf_path).stem + "_AMANCO_RELATORIO.xlsx")
        comparator = self.PriceComparator(excel_path, pdf_path, log_func)
        comparator.read_excel()
        comparator.read_pdf()
        df = comparator.merge_data()
        df = comparator.calculate_differences(df)
        comparator.save_result(df, str(out_path))
        return out_path


# ==============================================================================
# INTERFACE GRÁFICA (Atualizada com Cores e Logo)
# ==============================================================================

class ConstrumilGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("SISTEMA CONSTRUMIL - Conferência v6.0")
        self.root.geometry("900x750")

        # --- NOVAS CORES ---
        self.col_main = "#006cb5"  # Azul maior (Fundo principal)
        self.col_panel = "#91d8f6"  # Azul menor (Fundo dos painéis)
        self.col_accent1 = "#ec3237"  # Detalhes 1 (Vermelho - Botões/Abas)
        self.col_text = "#fefefe"  # Detalhes 2 (Texto branco)
        self.col_label_text = "#0A1931"  # Cor escura para texto sobre o painel claro

        self.root.configure(bg=self.col_main)

        # --- HEADER COM LOGO ---
        hdr = tk.Frame(root, bg=self.col_main)
        hdr.pack(fill="x", pady=(20, 10))

        try:
            # Tenta carregar a logo
            # Redimensiona a imagem para uma altura razoável (ex: 80px) mantendo proporção
            pil_img = Image.open("logo.png")
            baseheight = 80
            hpercent = (baseheight / float(pil_img.size[1]))
            wsize = int((float(pil_img.size[0]) * float(hpercent)))
            pil_img = pil_img.resize((wsize, baseheight), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(pil_img)

            logo_label = tk.Label(hdr, image=self.logo_img, bg=self.col_main)
            logo_label.pack(pady=5)
        except FileNotFoundError:
            # Fallback se não achar a logo
            tk.Label(hdr, text="CONSTRUMIL (Logo não encontrado)", font=("Arial Black", 24), bg=self.col_main,
                     fg=self.col_text).pack()
        except Exception as e:
            tk.Label(hdr, text=f"Erro no Logo: {e}", bg=self.col_main, fg=self.col_text).pack()

        tk.Label(hdr, text="", font=("Arial", 11), bg=self.col_main,
                 fg=self.col_text).pack()
        tk.Frame(root, bg=self.col_accent1, height=5).pack(fill="x", pady=(10, 0))

        # --- ESTILO DAS ABAS ---
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background=self.col_main, borderwidth=0)
        # Aba inativa (Cinza escuro para contraste com o azul principal)
        style.configure("TNotebook.Tab", background="#444", foreground=self.col_text, padding=[20, 10],
                        font=('Arial', 10, 'bold'))
        # Aba ativa (Vermelho acento)
        style.map("TNotebook.Tab", background=[("selected", self.col_accent1)],
                  foreground=[("selected", self.col_text)])

        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=20, pady=20)

        self.tab_deca = tk.Frame(self.nb, bg=self.col_main)
        self.tab_amanco = tk.Frame(self.nb, bg=self.col_main)

        self.nb.add(self.tab_deca, text="  DECA  ")
        self.nb.add(self.tab_amanco, text=" AMANCO ")

        self.setup_ui(self.tab_deca, "DECA", "txt")
        self.setup_ui(self.tab_amanco, "AMANCO", "pdf")

        tk.Label(root, text="Construmil © 2025", bg=self.col_main, fg=self.col_text).pack(pady=5)

    def setup_ui(self, parent, marca, ext):
        # Painel com a cor "Azul menor" (#91d8f6)
        card = tk.Frame(parent, bg=self.col_panel, padx=20, pady=20)
        card.pack(fill="both", expand=True)

        setattr(self, f"in_{marca}", tk.StringVar())
        setattr(self, f"xls_{marca}", tk.StringVar())

        lbl = "Arquivo TXT (Extraído):" if ext == "txt" else "Arquivo PDF (Original):"
        self.mk_input(card, lbl, getattr(self, f"in_{marca}"), ext)
        self.mk_input(card, "Planilha Excel:", getattr(self, f"xls_{marca}"), "xls")

        # Botão com a cor "Detalhes 1" (#ec3237)
        btn = tk.Button(card, text=f"PROCESSAR {marca}", bg=self.col_accent1, fg=self.col_text,
                        font=("Arial", 12, "bold"), relief="flat", cursor="hand2", activebackground="#c42b30",
                        command=lambda: self.run_process(marca))
        btn.pack(fill="x", pady=20, ipady=7)

        # Texto do log com cor escura para contraste no painel claro
        tk.Label(card, text="Log de Processamento:", bg=self.col_panel, fg=self.col_label_text,
                 font=("Arial", 9, "bold")).pack(anchor="w")

        # Área de log (Mantive fundo escuro para estilo "terminal")
        log = scrolledtext.ScrolledText(card, height=10, bg="#0A1931", fg="#00FF00", font=("Consolas", 9),
                                        relief="flat", bd=2)
        log.pack(fill="both", expand=True, pady=5)
        setattr(self, f"log_{marca}", log)

    def mk_input(self, parent, text, var, ext):
        f = tk.Frame(parent, bg=self.col_panel)
        f.pack(fill="x", pady=7)
        # Labels com cor escura para contraste
        tk.Label(f, text=text, bg=self.col_panel, fg=self.col_label_text, width=25, anchor="w",
                 font=("Arial", 10)).pack(side="left")
        entry = tk.Entry(f, textvariable=var, relief="flat", bd=2, font=("Arial", 10))
        entry.pack(side="left", fill="x", expand=True, padx=5, ipady=3)
        # Botão de arquivo com cor de acento
        tk.Button(f, text="📂 Procurar", command=lambda: self.browse(var, ext), bg=self.col_accent1, fg=self.col_text,
                  relief="flat", cursor="hand2").pack(side="left", padx=(5, 0))

    def browse(self, var, ext):
        ft = [("PDF", "*.pdf")] if ext == "pdf" else [("Text", "*.txt")] if ext == "txt" else [
            ("Excel", "*.xlsx *.xls")]
        f = filedialog.askopenfilename(filetypes=ft)
        if f: var.set(f)

    def log(self, marca, msg):
        l = getattr(self, f"log_{marca}")
        l.insert(tk.END, f"> {msg}\n")
        l.see(tk.END)

    def run_process(self, marca):
        inp = getattr(self, f"in_{marca}").get()
        xls = getattr(self, f"xls_{marca}").get()
        if not inp or not xls:
            messagebox.showwarning("Erro", "Selecione ambos os arquivos!")
            return
        self.log(marca, f"Iniciando processamento {marca}...")
        threading.Thread(target=lambda: self.worker(marca, inp, xls), daemon=True).start()

    def worker(self, marca, inp, xls):
        try:
            out_path = ""
            if marca == "DECA":
                engine = DecaEngine()
                out_path = engine.processar(inp, xls, lambda m: self.log(marca, m))
            else:
                engine = AmancoEngine()
                out_path = engine.processar(inp, xls, lambda m: self.log(marca, m))

            self.log(marca, f"SUCESSO! Relatório salvo.")
            messagebox.showinfo("Sucesso", f"Processamento concluído!\n\nSalvo em:\n{out_path}")

        except Exception as e:
            self.log(marca, f"ERRO FATAL: {e}")
            messagebox.showerror("Erro", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    ConstrumilGUI(root)
    root.mainloop()